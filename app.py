from flask import Flask, request, render_template, send_file, session
import pandas as pd
import os
import tempfile
import openpyxl
from openpyxl.styles import Font, Border
from scheduler import space_runs_min_gap_hard, find_max_feasible_gap
import logging
import traceback

app = Flask(__name__)
app.secret_key = "super-secret-key"

logging.basicConfig(level=logging.DEBUG)

@app.route("/", methods=["GET", "POST"])
def index():
    if request.method == "POST":
        try:
            file = request.files.get("file")
            if not file:
                return render_template("index.html", error="Please upload a file.")

            logging.debug(f"Received file of size: {file.content_length} bytes")
            logging.debug(f"Request headers: {dict(request.headers)}")

            all_sheets = pd.read_excel(file, sheet_name=None)
            processed_sheets = {}
            failed_sheets = []
            gap_info = []

            max_gap_setting = 8
            time_limit_per_try = 10

            for sheet_name, df in all_sheets.items():
                try:
                    max_gap = find_max_feasible_gap(df, max_gap=max_gap_setting, min_gap=1, time_limit=time_limit_per_try)
                    processed_df = space_runs_min_gap_hard(df, min_gap=max_gap)
                    if processed_df is None:
                        failed_sheets.append(sheet_name)
                    else:
                        processed_sheets[sheet_name] = processed_df

                        human_gaps = processed_df["Last Human Run"].dropna().astype(int)
                        dog_gaps = processed_df["Last Dog Run"].dropna().astype(int)
                        gap_info.append({
                            "sheet": sheet_name,
                            "human_gap": human_gaps.min() if not human_gaps.empty else "N/A",
                            "dog_gap": dog_gaps.min() if not dog_gaps.empty else "N/A"
                        })
                except Exception as e:
                    logging.error(f"Error processing sheet '{sheet_name}': {traceback.format_exc()}")
                    failed_sheets.append(sheet_name)

            if not processed_sheets:
                return render_template("index.html", error="Scheduling failed for all sheets. Try adjusting your data or settings.")

            temp_dir = tempfile.gettempdir()
            output_path = os.path.join(temp_dir, "processed_schedule.xlsx")

            with pd.ExcelWriter(output_path, engine='openpyxl') as writer:
                for sheet_name, df in processed_sheets.items():
                    df.to_excel(writer, sheet_name=sheet_name, index=True)

            wb = openpyxl.load_workbook(output_path)
            for sheet_name in wb.sheetnames:
                ws = wb[sheet_name]
                header_cell = ws['A1']
                header_cell.font = Font(bold=True)
                for row in range(2, ws.max_row + 1):
                    ws[f'A{row}'].font = Font(bold=False)

                no_border = Border()
                for row in ws.iter_rows():
                    for cell in row:
                        cell.border = no_border
            wb.save(output_path)

            original_filename = file.filename or "processed_schedule.xlsx"
            name_part, ext_part = os.path.splitext(original_filename)
            download_filename = f"{name_part}_sorted{ext_part}"

            session["download_filename"] = download_filename

            return render_template(
                "index.html",
                gap_info=gap_info,
                failed_sheets=failed_sheets,
                success=True,
                download_filename=download_filename,
            )

        except Exception as e:
            logging.error(f"Exception during scheduling: {traceback.format_exc()}")
            return render_template("index.html", error=f"Unexpected error: {e}")

    return render_template("index.html")


@app.route("/download")
def download_file():
    temp_dir = tempfile.gettempdir()
    output_path = os.path.join(temp_dir, "processed_schedule.xlsx")
    download_name = session.get("download_filename", "processed_schedule.xlsx")
    return send_file(
        output_path,
        as_attachment=True,
        download_name=download_name,
        mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        cache_timeout=0
    )


# New test endpoint to verify file uploads from mobile
@app.route("/test-upload", methods=["POST"])
def test_upload():
    file = request.files.get("file")
    if not file:
        return "No file received", 400
    size = file.content_length or len(file.read())
    logging.debug(f"/test-upload received file of size {size} bytes")
    return f"Received file of size {size} bytes"

if __name__ == "__main__":
    app.run(host="0.0.0.0", port=10000)
